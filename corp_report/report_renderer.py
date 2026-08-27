from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import FactPack, FinancialFact


NAVY = "002060"
LIGHT_BLUE = "D9EAF7"
GREY = "F2F2F2"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7C3D0")
HAIR = Side(style="hair", color="A6A6A6")
TITLE = Font(name="맑은 고딕", size=18, bold=True)
SECTION = Font(name="맑은 고딕", size=12, bold=True, color=WHITE)
HEADER = Font(name="맑은 고딕", size=10, bold=True)
BODY = Font(name="맑은 고딕", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
NUMBER_FORMAT = '#,##0.0;[Red](#,##0.0);-'
PERCENT_FORMAT = '0.0%;[Red](0.0%);-'


def _basis_label(fs_div: str) -> str:
    return "별도" if fs_div == "OFS" else "연결"


def _display_market_cap(value: object, date_value: object) -> str:
    if not isinstance(value, (int, float)) or value <= 0:
        return "확인 필요"
    return f"약 {value / 1_000_000_000_000:.2f}조원 ({date_value or '기준일 확인 필요'})"


def _value_for(pack: FactPack, period: str, account: str) -> float | None:
    """Return a displayed fact value in 억원 without reading an Excel formula back."""
    values = [fact.value_eok for fact in pack.financial_facts if fact.period_label == period and fact.standard_account == account and fact.value_eok is not None]
    if not values:
        return None
    return sum(values) if account == "차입금" else values[0]


def _ratio(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator <= 0:
        return None
    return numerator / denominator


def _format_amount(value: float | None) -> str:
    return "N/A" if value is None else f"{value:,.0f}억원"


def _format_ratio(value: float | None) -> str:
    return "N.M." if value is None else f"{value:.1%}"


def _growth_label(previous: float | None, current: float | None, profit_metric: bool = False) -> str:
    if previous is None or current is None:
        return "N/A"
    if profit_metric:
        if previous < 0 < current:
            return "흑자전환"
        if previous > 0 > current:
            return "적자전환"
        if previous < 0 and current < 0:
            return "적자축소" if current > previous else "적자확대"
    if previous <= 0:
        return "N.M."
    return f"{current / previous - 1:+.1%}"


def _latest_annual_periods(pack: FactPack) -> tuple[str | None, str | None]:
    annuals = [period for period in _periods(pack) if period.isdigit()]
    return (annuals[-1], annuals[-2]) if len(annuals) >= 2 else (annuals[-1], None) if annuals else (None, None)


def _financial_commentary(pack: FactPack) -> tuple[str, str, str]:
    latest_annual, prior_annual = _latest_annual_periods(pack)
    if not latest_annual:
        return ("성장성: 최근 연간 재무 수치 확인 필요", "수익성: 최근 연간 재무 수치 확인 필요", "안정성: 최근 연간 재무 수치 확인 필요")
    revenue = _value_for(pack, latest_annual, "매출액")
    operating = _value_for(pack, latest_annual, "영업이익")
    net_income = _value_for(pack, latest_annual, "당기순이익")
    prior_revenue = _value_for(pack, prior_annual, "매출액") if prior_annual else None
    prior_operating = _value_for(pack, prior_annual, "영업이익") if prior_annual else None
    assets = _value_for(pack, latest_annual, "자산총계")
    liability = _value_for(pack, latest_annual, "부채총계")
    equity = _value_for(pack, latest_annual, "자본총계")
    current_assets = _value_for(pack, latest_annual, "유동자산")
    current_liability = _value_for(pack, latest_annual, "유동부채")
    borrowings = _value_for(pack, latest_annual, "차입금")
    cash = _value_for(pack, latest_annual, "현금및현금성자산")
    interest = _value_for(pack, latest_annual, "이자비용")
    net_debt = borrowings - cash if borrowings is not None and cash is not None else None
    coverage = _ratio(operating, interest)
    growth = f"{latest_annual}년 매출액 {_format_amount(revenue)}(전년 대비 {_growth_label(prior_revenue, revenue)}), 영업이익 {_format_amount(operating)}({_growth_label(prior_operating, operating, True)}) 수준"
    profitability = f"{latest_annual}년 영업이익률 {_format_ratio(_ratio(operating, revenue))}, 순이익률 {_format_ratio(_ratio(net_income, revenue))} 수준"
    stability = f"{latest_annual}년 말 부채비율 {_format_ratio(_ratio(liability, equity))}, 유동비율 {_format_ratio(_ratio(current_assets, current_liability))}, 순차입금 {_format_amount(net_debt)}, 이자보상배율 {_format_ratio(coverage) if coverage is None else f'{coverage:.1f}배'} 수준"
    return growth, profitability, stability


def _short_text(value: object, limit: int = 120) -> str:
    text = " ".join(str(value or "").split())
    return text if len(text) <= limit else f"{text[:limit - 1]}…"


def _report_tone(value: object) -> str:
    """Use compact, nominal report prose on the executive-summary sheet only."""
    text = " ".join(str(value or "").split())
    replacements = (
        ("을 향해 나아갑니다.", " 도약 추진."),
        ("나아갑니다.", "추진."),
        ("하였습니다.", "함."), ("했습니다.", "함."), ("되었습니다.", "됨."),
        ("필요합니다.", "필요."), ("확인해야 합니다.", "확인 필요."),
        ("기재되어 있습니다.", "기재됨."), ("확인되었습니다.", "확인됨."),
        ("있습니다.", "있음."), ("없습니다.", "없음."),
        ("됩니다.", "됨."), ("입니다.", "임."), ("합니다.", "함."),
        ("나아갑니다", "추진"), ("하였습니다", "함"), ("했습니다", "함"), ("되었습니다", "됨"),
        ("필요합니다", "필요"), ("확인해야 합니다", "확인 필요"),
        ("기재되어 있습니다", "기재됨"), ("확인되었습니다", "확인됨"),
        ("있습니다", "있음"), ("없습니다", "없음"),
        ("됩니다", "됨"), ("입니다", "임"), ("합니다", "함"),
    )
    for original, replacement in replacements:
        text = text.replace(original, replacement)
    return text


def _source_text(title: object, disclosure_date: object = "") -> str:
    label = " ".join(str(title or "출처 확인 필요").split())
    date_text = str(disclosure_date or "").strip()
    return _short_text(f"{label} {date_text}".strip(), 38)


def _latest_document_source(pack: FactPack) -> tuple[str, str]:
    if not pack.documents:
        return "출처 확인 필요", ""
    document = sorted(pack.documents, key=lambda item: str(item.disclosure_date), reverse=True)[0]
    return _source_text(document.title, document.disclosure_date), document.url


def _corporate_source(pack: FactPack) -> tuple[str, str]:
    preferred_categories = {"사업의 내용", "회사 연혁", "회사 개요·감사보고서", "공식 홈페이지 사업소개", "공식 사업소개"}
    source = next((item for item in pack.corporate_sources if item.category in preferred_categories), None)
    if source:
        return _source_text(source.source_title, source.disclosure_date), source.url
    return _latest_document_source(pack)


def _financial_source(pack: FactPack, periods: list[str], account: str) -> tuple[str, str]:
    document_index = {item.document_id: item for item in pack.documents}
    display_ids = _display_fact_ids(pack.financial_facts)
    facts = [
        fact for fact in pack.financial_facts
        if fact.fact_id in display_ids and fact.period_label in periods and fact.standard_account == account
    ]
    documents = [document_index[fact.source_document_id] for fact in facts if fact.source_document_id in document_index]
    if not documents:
        return "재무 data 참조", "#'재무 data'!A1"
    latest = sorted(documents, key=lambda item: str(item.disclosure_date), reverse=True)[0]
    return _source_text(f"재무 data · {latest.title}", latest.disclosure_date), latest.url


def _style_chart_line(chart: LineChart, color: str = "0070C0") -> None:
    """Keep dense daily-price lines readable after A4 print scaling."""
    if not chart.series:
        return
    line = chart.series[0].graphicalProperties.line
    line.solidFill = color
    line.prstDash = "solid"
    line.width = 38100  # 3pt
    chart.series[0].marker.symbol = "none"


def _safe_sheet_name(name: str) -> str:
    return name.replace("[", "_").replace("]", "_").replace("/", "_")[:31]


def _style_section(ws, row: int, start_col: int, end_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, text)
    cell.font = SECTION
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 21


def _apply_table_header(ws, row: int, columns: list[str]) -> None:
    for col, label in enumerate(columns, start=1):
        cell = ws.cell(row, col, label)
        cell.font = HEADER
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = CENTER


def _write_data_sheet(ws, pack: FactPack) -> None:
    headers = [
        "회사명", "종목코드", "기준연도/분기", "보고서유형", "기간시작", "기간종료", "누적여부", "연결·별도",
        "계정과목 표준명", "원문 계정과목명", "계정ID", "재무제표", "시점·기간", "값(원)", "통화", "원 단위",
        "표시 단위", "값 구분", "출처 문서명", "공시일", "URL", "페이지 또는 표 위치", "표시 사용",
    ]
    ws.title = "재무 data"
    _apply_table_header(ws, 1, headers)
    doc_index = {document.document_id: document for document in pack.documents}
    display_fact_ids = _display_fact_ids(pack.financial_facts)
    for row_index, fact in enumerate(pack.financial_facts, start=2):
        document = doc_index.get(fact.source_document_id)
        values = [
            fact.company_name, fact.stock_code, fact.period_label, fact.report_type, fact.period_start, fact.period_end,
            "Y" if fact.is_cumulative else "N", _basis_label(fact.fs_div), fact.standard_account, fact.source_account,
            fact.account_id, fact.statement, fact.stock_or_flow, fact.value_krw, fact.currency, "원", "억원", fact.value_type,
            document.title if document else "확인 필요", document.disclosure_date if document else "", document.url if document else "", fact.source_location,
            "Y" if fact.fact_id in display_fact_ids else "N",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col, value)
            cell.font = BODY
            cell.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            cell.alignment = LEFT_WRAP if col in {1, 9, 10, 19, 21, 22} else CENTER
            if col == 14 and isinstance(value, (int, float)):
                cell.number_format = '#,##0;[Red](#,##0);-'
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:W{max(ws.max_row, 2)}"
    widths = [18, 11, 12, 12, 12, 12, 10, 10, 18, 28, 28, 12, 10, 16, 10, 10, 10, 12, 28, 12, 48, 30, 10]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _display_fact_ids(facts: Iterable[FinancialFact]) -> set[str]:
    """Select one reported total per metric; borrowing components remain additive."""
    selected: set[str] = set()
    selected_keys: set[tuple[str, str, str]] = set()
    aggregate_accounts = {"차입금"}
    preferred_statements = {
        "매출액": "PL", "매출원가": "PL", "영업이익": "PL", "당기순이익": "PL", "이자비용": "PL",
        "영업활동현금흐름": "CF",
        "자산총계": "BS", "유동자산": "BS", "부채총계": "BS", "유동부채": "BS", "자본총계": "BS",
        "현금및현금성자산": "BS", "매출채권": "BS", "재고자산": "BS", "자본금": "BS",
    }
    for fact in facts:
        if fact.value_krw is None:
            continue
        if fact.standard_account in aggregate_accounts:
            selected.add(fact.fact_id)
            continue
        key = (fact.period_label, fact.fs_div, fact.standard_account)
        preferred_statement = preferred_statements.get(fact.standard_account)
        if key not in selected_keys and (not preferred_statement or fact.statement == preferred_statement):
            selected.add(fact.fact_id)
            selected_keys.add(key)
    # Some reporters classify a line differently from the standard statement
    # code.  Retain one source fact as a fallback rather than rendering N/A.
    for fact in facts:
        if fact.value_krw is None or fact.standard_account in aggregate_accounts:
            continue
        key = (fact.period_label, fact.fs_div, fact.standard_account)
        if key not in selected_keys:
            selected.add(fact.fact_id)
            selected_keys.add(key)
    return selected


def _periods(pack: FactPack) -> list[str]:
    labels = []
    for fact in pack.financial_facts:
        if fact.period_label not in labels:
            labels.append(fact.period_label)
    annuals = sorted([label for label in labels if label.isdigit()], key=int)
    interims = [label for label in labels if not label.isdigit()]
    return annuals + interims


def _amount_formula(period_cell: str, account_cell: str, basis: str) -> str:
    criteria = f"'재무 data'!$C:$C,{period_cell},'재무 data'!$H:$H,\"{basis}\",'재무 data'!$I:$I,{account_cell},'재무 data'!$W:$W,\"Y\""
    return f'=IF(COUNTIFS({criteria})=0,"N/A",SUMIFS(\'재무 data\'!$N:$N,\'재무 data\'!$C:$C,{period_cell},\'재무 data\'!$H:$H,"{basis}",\'재무 data\'!$I:$I,{account_cell},\'재무 data\'!$W:$W,"Y")/100000000)'


def _write_financial_sheet(ws, pack: FactPack) -> None:
    basis = _basis_label(pack.reporting_policy["primary_fs_basis"])
    periods = _periods(pack)
    ws.title = "재무"
    ws["B2"] = "(단위 : 억원)"
    ws["B2"].font = Font(name="Arial", size=9)
    ws["B3"] = f"재무제표 기준: {basis}"
    ws["B3"].font = HEADER
    ws["B4"] = "구분"
    ws["B4"].font = HEADER
    ws["B4"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["B4"].alignment = CENTER
    ws["B4"].border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for index, label in enumerate(periods):
        amount_col = 3 + index * 2
        for col, header in ((amount_col, label), (amount_col + 1, "%")):
            cell = ws.cell(4, col, header)
            cell.font = HEADER
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            cell.alignment = CENTER
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            ws.column_dimensions[get_column_letter(col)].width = 14 if col == amount_col else 10

    rows = [
        (6, "매출액"), (7, "매출원가"), (8, "영업이익"), (9, "당기순이익"),
        (11, "자산총계"), (12, "유동자산"), (13, "현금및현금성자산"), (14, "매출채권"), (15, "재고자산"),
        (17, "부채총계"), (18, "유동부채"), (19, "자본총계"), (20, "자본금"),
        (21, "영업활동현금흐름"), (22, "이자비용"), (23, "차입금"),
    ]
    major_rows = {"매출액", "영업이익", "당기순이익", "자산총계", "부채총계", "자본총계"}
    composition_denominator = {
        6: 6, 7: 6, 8: 6, 9: 6, 11: 11, 12: 11, 13: 11, 14: 11,
        15: 11, 17: 11, 18: 17, 19: 11, 20: 11, 21: 6, 22: 6, 23: 11,
    }
    for row, label in rows:
        ws.cell(row, 2, label).font = HEADER if label in major_rows else BODY
        for index, _ in enumerate(periods):
            amount_col = 3 + index * 2
            amount_letter = get_column_letter(amount_col)
            cell = ws.cell(row, amount_col)
            cell.value = _amount_formula(f"{amount_letter}$4", f"$B{row}", basis)
            cell.number_format = NUMBER_FORMAT
            cell.font = Font(name="Arial", size=10, color="008000")
            cell.alignment = CENTER
            cell.border = Border(bottom=HAIR)
            percent = ws.cell(row, amount_col + 1)
            denominator_row = composition_denominator[row]
            percent.value = f'=IF(OR(NOT(ISNUMBER({amount_letter}{row})),NOT(ISNUMBER({amount_letter}{denominator_row})),{amount_letter}{denominator_row}=0),"N.M.",{amount_letter}{row}/{amount_letter}{denominator_row})'
            percent.number_format = PERCENT_FORMAT
            percent.alignment = CENTER
            percent.border = Border(bottom=HAIR)
    for row, label in [(26, "성장성"), (30, "수익성"), (34, "안정성"), (40, "활동성·현금흐름")]:
        ws.cell(row, 2, label).font = HEADER
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=GREY)
    ratio_rows = [
        (27, "매출액증감률", 6, False), (28, "영업이익증감률", 8, True),
        (31, "영업이익률", 8, None), (32, "순이익률", 9, None),
        (35, "부채비율", 17, "equity"), (36, "유동비율", 12, "current_liability"),
        (37, "순차입금", 23, "net_debt"), (38, "이자보상배율", 8, "coverage"),
        (41, "재고자산회전율", 7, "inventory_turnover"), (42, "재고자산회전일수", 7, "inventory_days"),
        (43, "영업활동현금흐름/매출", 21, None),
    ]
    for row, label, source_row, kind in ratio_rows:
        ws.cell(row, 2, label).font = BODY
        for index, _ in enumerate(periods):
            col = 3 + index * 2
            letter = get_column_letter(col)
            cell = ws.cell(row, col)
            if kind is False:
                prev = get_column_letter(col - 2)
                cell.value = "=\"N/A\"" if index == 0 else f'=IF(OR(RIGHT({letter}$4,1)="H",RIGHT({letter}$4,1)="Q"),"N/A",IF(OR(NOT(ISNUMBER({prev}{source_row})),NOT(ISNUMBER({letter}{source_row}))),"N/A",IF({prev}{source_row}<=0,"N.M.",{letter}{source_row}/{prev}{source_row}-1)))'
                cell.number_format = PERCENT_FORMAT
            elif kind is True:
                prev = get_column_letter(col - 2)
                cell.value = "=\"N/A\"" if index == 0 else f'=IF(OR(RIGHT({letter}$4,1)="H",RIGHT({letter}$4,1)="Q"),"N/A",IF(OR(NOT(ISNUMBER({prev}{source_row})),NOT(ISNUMBER({letter}{source_row}))),"N/A",IF(AND({prev}{source_row}<0,{letter}{source_row}>0),"흑자전환",IF(AND({prev}{source_row}>0,{letter}{source_row}<0),"적자전환",IF(AND({prev}{source_row}<0,{letter}{source_row}<0),IF({letter}{source_row}>{prev}{source_row},"적자축소","적자확대"),IF({prev}{source_row}<=0,"N.M.",{letter}{source_row}/{prev}{source_row}-1))))))'
                cell.number_format = PERCENT_FORMAT
            elif kind is None:
                cell.value = f'=IF(OR(NOT(ISNUMBER({letter}{source_row})),NOT(ISNUMBER({letter}6)),{letter}6<=0),"N.M.",{letter}{source_row}/{letter}6)'
                cell.number_format = PERCENT_FORMAT
            elif kind == "equity":
                cell.value = f'=IF(OR(NOT(ISNUMBER({letter}17)),NOT(ISNUMBER({letter}19)),{letter}19<=0),"N.M.",{letter}17/{letter}19)'
                cell.number_format = PERCENT_FORMAT
            elif kind == "current_liability":
                cell.value = f'=IF(OR(NOT(ISNUMBER({letter}12)),NOT(ISNUMBER({letter}18)),{letter}18<=0),"N.M.",{letter}12/{letter}18)'
                cell.number_format = PERCENT_FORMAT
            elif kind == "net_debt":
                cell.value = f'=IF(OR(NOT(ISNUMBER({letter}23)),NOT(ISNUMBER({letter}13))),"N/A",{letter}23-{letter}13)'
                cell.number_format = NUMBER_FORMAT
            elif kind == "inventory_turnover":
                prev = get_column_letter(col - 2)
                cell.value = "=\"N/A\"" if index == 0 else f'=IF(OR(NOT(ISNUMBER({letter}7)),NOT(ISNUMBER({letter}15)),NOT(ISNUMBER({prev}15)),({letter}15+{prev}15)<=0),"N.M.",{letter}7/AVERAGE({prev}15,{letter}15))'
                cell.number_format = '0.0x;[Red](0.0x);-'
            elif kind == "inventory_days":
                cell.value = f'=IF(OR(NOT(ISNUMBER({letter}41)),{letter}41<=0),"N.M.",365/{letter}41)'
                cell.number_format = '0.0;[Red](0.0);-'
            else:
                cell.value = f'=IF(OR(NOT(ISNUMBER({letter}8)),NOT(ISNUMBER({letter}22)),{letter}22<=0),"N.M.",{letter}8/{letter}22)'
                cell.number_format = '0.0x;[Red](0.0x);-'
            cell.alignment = CENTER
            cell.border = Border(bottom=HAIR)
            percent = ws.cell(row, col + 1)
            percent.value = ""
            percent.border = Border(bottom=HAIR)
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "C5"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1


def _write_summary_sheet(ws, pack: FactPack, price_data_ws=None) -> None:
    company = str(pack.entity.get("company_name", "기업"))
    ws.title = "본장"
    ws.merge_cells("C1:AL1")
    ws["C1"] = f"◐ {company}"
    ws["C1"].font = TITLE
    ws["C3"] = f"< {datetime.now().strftime('%Y. %m. %d')} >"
    ws["C3"].alignment = CENTER
    _style_section(ws, 5, 4, 38, "가. 회사 및 사업개요")
    profile = pack.corporate_profile
    corporate_source, corporate_url = _corporate_source(pack)
    profile_fields = [
        ("상호", company), ("대표자", profile.get("ceo_name", "확인 필요")), ("소재지", profile.get("address", "확인 필요")),
        ("설립일자", profile.get("establishment_date", "확인 필요")), ("종목코드", pack.entity.get("stock_code") or "비상장"),
        ("기업형태", "상장" if pack.entity.get("listing_status") == "listed" else "비상장"),
        ("종속관계", f"{profile.get('parent_company')} {profile.get('parent_company_ratio'):.2f}%" if profile.get("parent_company") and isinstance(profile.get("parent_company_ratio"), (int, float)) else profile.get("parent_company", "확인 필요")),
        ("종업원 수", f"{profile.get('employee_count'):,}명" if isinstance(profile.get("employee_count"), int) else "확인 필요"),
        ("최대주주", f"{profile.get('largest_holder', '확인 필요')} {profile.get('largest_holder_ratio', 0):.2f}%" if isinstance(profile.get("largest_holder_ratio"), (int, float)) else "확인 필요"),
        ("시가총액", _display_market_cap(pack.entity.get("market_cap_krw"), pack.entity.get("market_price_date"))),
    ]

    def apply_source(row: int, source_title: str, source_url: str, source_start: int = 29, source_end: int = 32, link_start: int = 33, link_end: int = 38) -> None:
        ws.merge_cells(start_row=row, start_column=source_start, end_row=row, end_column=source_end)
        ws.merge_cells(start_row=row, start_column=link_start, end_row=row, end_column=link_end)
        source_cell = ws.cell(row, source_start, source_title or "출처 확인 필요")
        link_cell = ws.cell(row, link_start, "원문 보기" if source_url else "확인 필요")
        source_cell.font = Font(name="맑은 고딕", size=8, color="555555")
        link_cell.font = Font(name="맑은 고딕", size=8, color="0563C1", underline="single" if source_url else None)
        source_cell.alignment = LEFT_WRAP
        link_cell.alignment = CENTER
        if source_url:
            link_cell.hyperlink = source_url
        for cell in (source_cell, link_cell):
            cell.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

    def profile_field(row: int, position: tuple[int, int, int, int, int, int], label: str, value: object) -> None:
        label_start, label_end, value_start, value_end, source_start, link_start = position
        ws.merge_cells(start_row=row, start_column=label_start, end_row=row, end_column=label_end)
        ws.merge_cells(start_row=row, start_column=value_start, end_row=row, end_column=value_end)
        label_cell = ws.cell(row, label_start, label)
        value_cell = ws.cell(row, value_start, _report_tone(value))
        label_cell.fill = PatternFill("solid", fgColor=GREY)
        label_cell.font = HEADER
        value_cell.font = Font(name="맑은 고딕", size=9, bold=label in {"상호", "종목코드"})
        for cell in (label_cell, value_cell):
            cell.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            cell.alignment = LEFT_WRAP if cell is value_cell else CENTER
        apply_source(row, corporate_source, corporate_url, source_start, source_start + 3, link_start, link_start + 1)

    profile_positions = ((5, 7, 8, 14, 15, 19), (21, 23, 24, 30, 31, 35))
    for index, (label, value) in enumerate(profile_fields):
        row = 7 + index // 2
        profile_field(row, profile_positions[index % 2], label, value)
        if label == "소재지":
            ws.row_dimensions[row].height = 32

    def labeled_row(row: int, label: str, value: object, source_title: str = "", source_url: str = "", height: int = 23) -> None:
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=28)
        label_cell = ws.cell(row, 5, label)
        value_cell = ws.cell(row, 9, _report_tone(value))
        label_cell.fill = PatternFill("solid", fgColor=GREY)
        label_cell.font = HEADER
        value_cell.font = Font(name="맑은 고딕", size=9)
        for cell in (label_cell, value_cell):
            cell.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            cell.alignment = LEFT_WRAP if cell is value_cell else CENTER
        apply_source(row, source_title, source_url)
        ws.row_dimensions[row].height = height

    # Keep each company attribute, capability and timeline entry on its own row.
    # This preserves an executive-report layout instead of putting a paragraph in
    # a single merged cell.
    labeled_row(13, "주요 사업", str(profile.get("business_summary", "")).strip() or "확인 필요", corporate_source, corporate_url, height=42)
    labeled_row(14, "성장 전략", str(profile.get("growth_strategy", "")).strip() or "확인 필요", corporate_source, corporate_url, height=42)
    competencies = [str(value).strip() for value in profile.get("core_competencies", []) if str(value).strip()]
    for index in range(4):
        labeled_row(15 + index, "핵심 역량" if index == 0 else "", competencies[index] if index < len(competencies) else "", corporate_source, corporate_url)
    labeled_row(19, "대표이사 약력", str(profile.get("ceo_bio", "")).strip() or "확인 필요", corporate_source, corporate_url, height=30)

    ws.merge_cells("E20:H20")
    ws["E20"] = "핵심 연혁"
    ws["E20"].font = HEADER
    ws["E20"].fill = PatternFill("solid", fgColor=GREY)
    ws["E20"].alignment = CENTER
    chronology_source_index = {source.source_document_id: source for source in pack.corporate_sources}
    for index in range(6):
        row = 21 + index
        item = pack.chronology[index] if index < len(pack.chronology) else {}
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=28)
        ws.cell(row, 9, item.get("date", ""))
        ws.cell(row, 12, _report_tone(item.get("event", "")))
        for col in (5, 9, 12):
            cell = ws.cell(row, col)
            cell.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            cell.alignment = LEFT_WRAP if col == 12 else CENTER
            cell.font = BODY if col != 12 else Font(name="맑은 고딕", size=9)
        if index == 0:
            ws.cell(row, 5, "연혁")
            ws.cell(row, 5).fill = PatternFill("solid", fgColor=GREY)
            ws.cell(row, 5).font = HEADER
        source = next((chronology_source_index[source_id] for source_id in item.get("source_ids", []) if source_id in chronology_source_index), None)
        source_label, source_url = (_source_text(source.source_title, source.disclosure_date), source.url) if source else (corporate_source, corporate_url)
        apply_source(row, source_label, source_url)
        ws.row_dimensions[row].height = 22

    _style_section(ws, 28, 4, 38, "나. 재무 현황")
    periods = _periods(pack)[-5:]
    ws.merge_cells("E30:H30")
    ws["E30"] = "지표"
    ws["E30"].font = HEADER
    ws["E30"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["E30"].alignment = CENTER
    for period_index, period in enumerate(periods):
        start_col = 9 + period_index * 4
        end_col = start_col + 3
        ws.merge_cells(start_row=30, start_column=start_col, end_row=30, end_column=end_col)
        cell = ws.cell(30, start_col, period)
        cell.font = HEADER
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.alignment = CENTER
    for start_col, end_col, label in ((29, 32, "출처"), (33, 38, "링크")):
        ws.merge_cells(start_row=30, start_column=start_col, end_row=30, end_column=end_col)
        cell = ws.cell(30, start_col, label)
        cell.font = HEADER
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.alignment = CENTER
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    summary_rows = [
        (31, "매출액", 6, False), (32, "매출액증감률", 27, True),
        (33, "영업이익", 8, False), (34, "영업이익증감률", 28, True),
        (35, "당기순이익", 9, False), (36, "영업이익률", 31, True), (37, "순이익률", 32, True),
        (38, "자산총계", 11, False), (39, "부채총계", 17, False), (40, "자본총계", 19, False),
        (41, "부채비율", 35, True), (42, "유동비율", 36, True), (43, "순차입금", 37, False), (44, "이자보상배율", 38, False),
    ]
    for row, label, financial_row, is_percent in summary_rows:
        source_title, source_url = _financial_source(pack, periods, label.replace("증감률", "").replace("률", "") if label in {"매출액증감률", "영업이익증감률", "영업이익률", "순이익률"} else label)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        ws.cell(row, 5, label)
        for period_index, _ in enumerate(periods):
            start_col = 9 + period_index * 4
            end_col = start_col + 3
            source_col = get_column_letter(3 + (len(_periods(pack)) - len(periods) + period_index) * 2)
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            ws.cell(row, start_col, f"='재무'!{source_col}{financial_row}")
            ws.cell(row, start_col).number_format = PERCENT_FORMAT if is_percent else NUMBER_FORMAT
            ws.cell(row, start_col).font = Font(name="맑은 고딕", size=9, color="008000")
        for col in [5] + [9 + period_index * 4 for period_index in range(len(periods))]:
            ws.cell(row, col).border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            ws.cell(row, col).alignment = CENTER
            if row in {32, 34, 36, 37, 41, 42}:
                ws.cell(row, col).fill = PatternFill("solid", fgColor="F8FBFD")
        apply_source(row, source_title, source_url)
    growth_note, profitability_note, stability_note = _financial_commentary(pack)
    financial_note_source, financial_note_url = _financial_source(pack, periods, "매출액")
    labeled_row(45, "성장성 해석", growth_note, financial_note_source, financial_note_url)
    labeled_row(46, "수익성 해석", profitability_note, financial_note_source, financial_note_url)
    labeled_row(47, "안정성 해석", stability_note, financial_note_source, financial_note_url)

    _style_section(ws, 49, 4, 38, "다. 주요 사항")
    ws.merge_cells("E50:S50")
    issue_query = str(pack.reporting_policy.get("issue_query", "")).strip()
    ws["E50"] = f"검토 요청: {issue_query}" if issue_query else "검토 요청: N/A"
    ws["E50"].font = Font(name="맑은 고딕", size=9, italic=True, color="555555")
    ws["E50"].alignment = LEFT_WRAP
    issue_matters = [matter for matter in pack.major_matters if not matter.category.startswith("주가 변동")][:3]
    if issue_matters:
        for index, matter in enumerate(issue_matters):
            row = 52 + index * 2
            category = matter.category.removeprefix("Gemini 분석·").removeprefix("웹 이슈 조사·")
            source_title = _source_text(matter.source_title, matter.disclosure_date)
            labeled_row(row, f"◯ {category}", _short_text(matter.fact, 150), source_title, matter.url, height=34)
            labeled_row(row + 1, "시사점", _short_text(matter.interpretation, 150), source_title, matter.url, height=34)
    else:
        labeled_row(52, "◯ 주요 사항", "입력 이슈 없음. 웹 이슈 질의 입력 시 근거 기반 요약 표시", corporate_source, corporate_url)

    subsidiary_section_row = 60
    _style_section(ws, subsidiary_section_row, 4, 38, "라. 자회사 등 현황")
    subsidiary_header_row = subsidiary_section_row + 2
    subsidiary_headers = [(5, 8, "사업군"), (9, 14, "자회사명"), (15, 18, "지분율"), (19, 28, "사업영역·비고"), (29, 32, "출처"), (33, 38, "링크")]
    for start_col, end_col, label in subsidiary_headers:
        ws.merge_cells(start_row=subsidiary_header_row, start_column=start_col, end_row=subsidiary_header_row, end_column=end_col)
        cell = ws.cell(subsidiary_header_row, start_col, label)
        cell.font = HEADER
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.alignment = CENTER
        cell.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
    if pack.subsidiaries:
        for offset, subsidiary in enumerate(pack.subsidiaries[:5], start=0):
            row = subsidiary_header_row + 1 + offset
            values = [
                subsidiary.get("사업군", "확인 필요"),
                subsidiary.get("자회사명", "확인 필요"),
                subsidiary.get("지분율", "확인 필요"),
                " · ".join(value for value in [subsidiary.get("사업영역", ""), subsidiary.get("비고", "")] if value) or "확인 필요",
            ]
            for (start_col, end_col, _), value in zip(subsidiary_headers[:4], values):
                ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
                cell = ws.cell(row, start_col, _report_tone(value))
                cell.font = Font(name="맑은 고딕", size=9)
                cell.alignment = LEFT_WRAP
                cell.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            apply_source(row, _source_text(subsidiary.get("출처 문서", ""), subsidiary.get("출처일", "")), str(subsidiary.get("출처 URL", "")))
            ws.row_dimensions[row].height = 24
        subsidiary_last_row = subsidiary_header_row + min(5, len(pack.subsidiaries))
    else:
        row = subsidiary_header_row + 1
        labeled_row(row, "자회사", "최신 사업보고서 또는 감사보고서 주석 확인 필요", corporate_source, corporate_url)
        subsidiary_last_row = row

    price_section_row = subsidiary_last_row + 2
    price_end_row = price_section_row
    if pack.entity.get("listing_status") == "listed" and pack.price_history:
        _style_section(ws, price_section_row, 4, 38, "마. 주가 추이")
        points = sorted(pack.price_history, key=lambda item: item.trading_date)
        start, end = points[0], points[-1]
        return_1y = end.close / start.close - 1 if start.close else None
        low = min(points, key=lambda item: item.close)
        high = max(points, key=lambda item: item.close)
        price_source = "Yahoo Finance 일별 종가"
        price_url = start.source_url
        price_summary = (
            f"관측기간 {start.trading_date}~{end.trading_date}, 종가 {start.close:,.0f}원→{end.close:,.0f}원 "
            f"({_format_ratio(return_1y)}), 저점 {low.close:,.0f}원({low.trading_date}), 고점 {high.close:,.0f}원({high.trading_date})"
        )
        labeled_row(price_section_row + 2, "최근 1년 요약", price_summary, price_source, price_url, height=28)
        chart = LineChart()
        chart.x_axis.delete = True
        chart.y_axis.delete = True
        chart.height = 6.4
        chart.width = 18
        chart_source_ws = price_data_ws or ws
        data = Reference(chart_source_ws, min_col=2, min_row=1, max_row=1 + len(points))
        categories = Reference(chart_source_ws, min_col=1, min_row=2, max_row=1 + len(points))
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        chart.legend = None
        _style_chart_line(chart)
        ws.add_chart(chart, f"E{price_section_row + 4}")
        price_end_row = price_section_row + 25

    for column in range(1, 39):
        ws.column_dimensions[get_column_letter(column)].width = 3.2
    ws.column_dimensions["C"].width = 12
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.row_breaks.append(__import__("openpyxl").worksheet.pagebreak.Break(id=48))
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 2
    ws.page_margins.left = 0.22
    ws.page_margins.right = 0.22
    ws.page_margins.top = 0.28
    ws.page_margins.bottom = 0.28
    ws.print_options.horizontalCentered = True
    ws.print_area = f"D1:AL{price_end_row}"


def _write_price_data_sheet(ws, pack: FactPack) -> None:
    ws.title = "주가 data"
    headers = ["거래일", "종가", "거래량", "출처 URL", "값 구분"]
    _apply_table_header(ws, 1, headers)
    for row, point in enumerate(pack.price_history, start=2):
        ws.cell(row, 1, datetime.fromisoformat(point.trading_date).date())
        ws.cell(row, 2, point.close)
        ws.cell(row, 3, point.volume)
        ws.cell(row, 4, point.source_url)
        ws.cell(row, 5, "collected")
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        ws.cell(row, 2).number_format = '#,##0.00;[Red](#,##0.00);-'
        ws.cell(row, 3).number_format = '#,##0;[Red](#,##0);-'
        for col in range(1, 6):
            ws.cell(row, col).border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            ws.cell(row, col).alignment = LEFT_WRAP if col == 4 else CENTER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(ws.max_row, 2)}"
    for col, width in enumerate([14, 14, 18, 72, 14], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_price_sheet(ws, pack: FactPack) -> None:
    ws.title = "주가"
    ws["A1"] = "최근 1년 주가 분석"
    ws["A1"].font = TITLE
    ws["A2"] = "원천: 주가 data 시트의 Yahoo Finance 일별 종가"
    ws["A2"].font = Font(name="맑은 고딕", size=9, color="555555")
    ws.merge_cells("A2:F2")
    if not pack.price_history:
        ws.merge_cells("A4:F4")
        ws["A4"] = "상장사 가격 데이터를 수집하지 못했습니다."
        return

    points = sorted(pack.price_history, key=lambda item: item.trading_date)
    low = min(points, key=lambda item: item.close)
    high = max(points, key=lambda item: item.close)
    start, end = points[0], points[-1]
    return_1y = (end.close / start.close - 1) if start.close else None
    summary = [
        ("관측 시작일", datetime.fromisoformat(start.trading_date).date()),
        ("관측 종료일", datetime.fromisoformat(end.trading_date).date()),
        ("시작 종가", start.close),
        ("종료 종가", end.close),
        ("1년 변동률", return_1y),
        ("기간 최저 종가", f"{low.close:,.0f}원 ({low.trading_date})"),
        ("기간 최고 종가", f"{high.close:,.0f}원 ({high.trading_date})"),
        ("관측 거래일", len(points)),
    ]
    for row, (label, value) in enumerate(summary, start=4):
        ws.cell(row, 1, label).font = HEADER
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=GREY)
        ws.cell(row, 2, value).font = BODY
        for col in (1, 2):
            ws.cell(row, col).border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            ws.cell(row, col).alignment = CENTER
    ws["B4"].number_format = ws["B5"].number_format = "yyyy-mm-dd"
    ws["B6"].number_format = ws["B7"].number_format = '#,##0.00;[Red](#,##0.00);-'
    ws["B8"].number_format = PERCENT_FORMAT

    event_matters = [item for item in pack.major_matters if item.category == "주가 변동·이슈 대조"]
    ws["H2"] = "주요 변동 이벤트"
    ws["H2"].font = HEADER
    for col, label in enumerate(["거래일", "관련 이벤트", "출처"], start=8):
        cell = ws.cell(3, col, label)
        cell.font = HEADER
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = CENTER
    for row, matter in enumerate(event_matters[:2], start=4):
        values = [matter.disclosure_date, _short_text(matter.fact, 110), matter.url]
        for col, value in enumerate(values, start=8):
            ws.cell(row, col, value)
            ws.cell(row, col).alignment = LEFT_WRAP if col in {9, 10} else CENTER
            ws.cell(row, col).border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
    if not event_matters:
        ws.merge_cells("H4:J4")
        ws["H4"] = "수집된 이슈와 날짜가 가까운 급변 구간을 확인하지 못했습니다."

    headers = ["거래일", "종가", "거래량", "1일 변동률", "5일 변동률", "20일 변동률", "이슈 대조"]
    _apply_table_header(ws, 13, headers)
    event_text = {item.disclosure_date: item.fact for item in event_matters}
    for index, point in enumerate(points, start=0):
        row = 14 + index
        source_row = 2 + index
        ws.cell(row, 1, f"='주가 data'!A{source_row}")
        ws.cell(row, 2, f"='주가 data'!B{source_row}")
        ws.cell(row, 3, f"='주가 data'!C{source_row}")
        ws.cell(row, 4, '="N/A"' if index == 0 else f'=IF(OR(NOT(ISNUMBER(B{row})),NOT(ISNUMBER(B{row - 1})),B{row - 1}<=0),"N/A",B{row}/B{row - 1}-1)')
        ws.cell(row, 5, '="N/A"' if index < 5 else f'=IF(OR(NOT(ISNUMBER(B{row})),NOT(ISNUMBER(B{row - 5})),B{row - 5}<=0),"N/A",B{row}/B{row - 5}-1)')
        ws.cell(row, 6, '="N/A"' if index < 20 else f'=IF(OR(NOT(ISNUMBER(B{row})),NOT(ISNUMBER(B{row - 20})),B{row - 20}<=0),"N/A",B{row}/B{row - 20}-1)')
        ws.cell(row, 7, event_text.get(point.trading_date, ""))
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            cell.alignment = LEFT_WRAP if col == 7 else CENTER
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        ws.cell(row, 2).number_format = '#,##0.00;[Red](#,##0.00);-'
        ws.cell(row, 3).number_format = '#,##0;[Red](#,##0);-'
        for col in (4, 5, 6):
            ws.cell(row, col).number_format = PERCENT_FORMAT
        if point.trading_date in event_text:
            for col in range(1, 8):
                ws.cell(row, col).fill = PatternFill("solid", fgColor="FFF2CC")

    chart = LineChart()
    chart.x_axis.delete = True
    chart.y_axis.delete = True
    chart.height = 9
    chart.width = 16
    chart.add_data(Reference(ws, min_col=2, min_row=13, max_row=13 + len(points)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=14, max_row=13 + len(points)))
    chart.legend = None
    _style_chart_line(chart)
    ws.add_chart(chart, "H10")
    ws.freeze_panes = "A14"
    ws.auto_filter.ref = f"A13:G{13 + len(points)}"
    for col, width in enumerate([14, 14, 18, 14, 14, 14, 58, 14, 55, 55], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_subsidiary_sheet(ws, pack: FactPack) -> None:
    ws.title = "자회사 등"
    headers = ["사업군", "자회사명", "지분율", "사업영역", "소재지", "자산(공시 표기)", "매출액(공시 표기)", "당기순이익(공시 표기)", "주요 생산시설 또는 핵심 역량", "비고", "출처 문서", "출처일", "출처 위치", "출처 URL"]
    _apply_table_header(ws, 2, headers)
    if not pack.subsidiaries:
        ws.merge_cells("A3:N3")
        ws["A3"] = "자동 추출 대상 공시가 확보되지 않았습니다. 사업보고서 주석 확인 필요"
        ws["A3"].alignment = LEFT_WRAP
    else:
        for row, subsidiary in enumerate(sorted(pack.subsidiaries, key=lambda item: (item.get("사업군", ""), item.get("자회사명", ""))), start=3):
            for col, header in enumerate(headers, start=1):
                ws.cell(row, col, subsidiary.get(header, "확인 필요"))
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:N{max(ws.max_row, 3)}"
    for col, width in enumerate([18, 24, 10, 24, 18, 16, 16, 18, 44, 24, 32, 14, 34, 60], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_matters_sheet(ws, pack: FactPack) -> None:
    ws.title = "주요사항"
    headers = ["분류", "사실", "해석", "검증 상태", "출처 문서명", "공시일", "URL"]
    _apply_table_header(ws, 2, headers)
    all_matters = [*pack.major_matters, *pack.corporate_sources]
    for row, matter in enumerate(all_matters, start=3):
        values = [matter.category, matter.fact, matter.interpretation, matter.verification_status, matter.source_title, matter.disclosure_date, matter.url]
        for col, value in enumerate(values, start=1):
            ws.cell(row, col, value)
            ws.cell(row, col).alignment = LEFT_WRAP if col in {2, 3, 5, 7} else CENTER
            ws.cell(row, col).border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
    if not all_matters:
        ws.merge_cells("A3:G3")
        ws["A3"] = "입력한 이슈가 없습니다. 웹 이슈 질의를 입력하면 조사 결과와 근거 URL을 표시합니다."
    widths = [14, 38, 48, 14, 34, 14, 55]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def render_report(pack: FactPack, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    financial = workbook.create_sheet()
    financial_data = workbook.create_sheet()
    subsidiaries = workbook.create_sheet()
    matters = workbook.create_sheet()
    _write_data_sheet(financial_data, pack)
    _write_financial_sheet(financial, pack)
    price_data = None
    if pack.entity.get("listing_status") == "listed" and pack.price_history:
        price = workbook.create_sheet()
        price_data = workbook.create_sheet()
        _write_price_data_sheet(price_data, pack)
        _write_price_sheet(price, pack)
    _write_summary_sheet(summary, pack, price_data)
    _write_subsidiary_sheet(subsidiaries, pack)
    _write_matters_sheet(matters, pack)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    safe_name = str(pack.entity.get("company_name", "기업")).replace("/", "_")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{safe_name}_기업분석보고서_{stamp}.xlsx"
    workbook.save(output_path)
    (output_path.with_suffix(".factpack.json")).write_text(json.dumps(pack.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path
