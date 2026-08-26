import unittest
from math import nan
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl
import pandas as pd

from corp_report.models import FactPack, FinancialFact, MatterFact, PricePoint, SourceDocument
from corp_report.collector import DartFactPackCollector, _optional_int, standard_account
from corp_report.gemini_research import _to_matters
from corp_report.report_renderer import render_report
from corp_report.validation import growth_display, validate_fact_pack
from corp_report.web_research import company_context_matters, company_context_profile, price_event_matters


def _fact(account: str, value: int) -> FinancialFact:
    return FinancialFact(
        fact_id=account,
        company_name="테스트",
        stock_code="",
        fiscal_year=2025,
        period_label="2025",
        report_type="annual",
        period_start="2025-01-01",
        period_end="2025-12-31",
        is_cumulative=True,
        fs_div="OFS",
        statement="BS",
        stock_or_flow="stock",
        standard_account=account,
        source_account=account,
        account_id="",
        value_krw=value,
        currency="KRW",
        source_document_id="doc-1",
        source_location="table",
    )


class ValidationTests(unittest.TestCase):
    def _pack(self) -> FactPack:
        return FactPack(
            pack_id="test",
            generated_at="2026-01-01T00:00:00+09:00",
            entity={"company_name": "테스트", "listing_status": "unlisted", "stock_code": ""},
            reporting_policy={"primary_fs_basis": "OFS", "display_unit": "억원"},
            documents=[SourceDocument("doc-1", "사업보고서", "2026-03-01", "https://example.com")],
            financial_facts=[
                _fact("매출액", 100), _fact("영업이익", 20), _fact("당기순이익", 10),
                _fact("자산총계", 100), _fact("부채총계", 60), _fact("자본총계", 40),
            ],
            price_history=[PricePoint("2026-01-02", 100.0, 1_000, "https://example.com/price")],
        )

    def test_growth_labels(self) -> None:
        self.assertEqual(growth_display(-10, 5, profit_metric=True), "흑자전환")
        self.assertEqual(growth_display(5, -10, profit_metric=True), "적자전환")
        self.assertEqual(growth_display(0, 5), "N.M.")

    def test_nan_is_not_converted_to_integer(self) -> None:
        self.assertIsNone(_optional_int(nan))

    def test_cost_of_sales_is_not_classified_as_revenue(self) -> None:
        self.assertEqual(standard_account("매출원가", ""), "매출원가")

    def test_revenue_related_accounts_are_not_total_revenue(self) -> None:
        self.assertEqual(standard_account("매출채권및기타채권", "ifrs-full_TradeAndOtherCurrentReceivables"), "매출채권")
        self.assertEqual(standard_account("매출총이익", "ifrs-full_GrossProfit"), "매출총이익")
        self.assertEqual(standard_account("제품매출", "dart_RevenueFromSaleOfGoodsProduct"), "제품매출")
        self.assertEqual(standard_account("매출액", "ifrs-full_Revenue"), "매출액")

    def test_audit_html_match_accepts_prefixed_total_revenue_only(self) -> None:
        frame = pd.DataFrame([
            {"재무항목": "매출채권", "당기금액": 100, "matched_exact": False, "테이블번호": 1, "행번호": 1},
            {"재무항목": "I. 매출", "당기금액": 500, "matched_exact": False, "테이블번호": 2, "행번호": 3},
            {"재무항목": "III. 매출총이익", "당기금액": 200, "matched_exact": False, "테이블번호": 2, "행번호": 5},
        ])
        matches = DartFactPackCollector._audit_matches(frame, "매출액")
        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["당기금액"], 500)

    def test_donga_audit_context_is_source_linked(self) -> None:
        profile = company_context_profile("동아제약")
        matters = company_context_matters("동아제약")
        self.assertIn("박카스", profile["business_summary"])
        self.assertEqual(len(matters), 1)
        self.assertEqual(matters[0].verification_status, "needs_review")
        self.assertTrue(matters[0].url.startswith("https://"))

    def test_audit_overview_profile_uses_only_explicit_purpose_clause(self) -> None:
        source = MatterFact(
            "회사 개요·감사보고서 주석",
            "당사는 원료의약품의 제조, 판매 및 수출입업 등을 목적사업으로 영위하고 있습니다.",
            "감사보고서 주석의 일반사항 원문 발췌",
            "verified",
            "audit-overview",
            "감사보고서",
            "2026-03-01",
            "https://example.com/audit",
        )
        profile = DartFactPackCollector._audit_overview_profile([source])
        self.assertIn("원료의약품", profile["business_summary"])

    def test_balance_sheet_tie_out(self) -> None:
        results = validate_fact_pack(self._pack())
        self.assertFalse([result for result in results if result["rule"] == "balance_sheet_tie_out"])

    def test_renderer_links_finance_to_raw_data(self) -> None:
        with TemporaryDirectory() as temp_dir:
            output = render_report(self._pack(), Path(temp_dir))
            workbook = openpyxl.load_workbook(output, data_only=False)
            self.assertEqual(workbook.sheetnames, ["본장", "재무", "재무 data", "자회사 등", "주요사항"])
            self.assertTrue(workbook["재무"]["C6"].value.startswith("=IF(COUNTIFS"))
            self.assertTrue(workbook["본장"]["I31"].value.startswith("='재무'!"))

    def test_renderer_adds_price_sheets_only_for_listed_company(self) -> None:
        with TemporaryDirectory() as temp_dir:
            pack = self._pack()
            pack.entity.update({"listing_status": "listed", "stock_code": "005930"})
            output = render_report(pack, Path(temp_dir))
            workbook = openpyxl.load_workbook(output, data_only=False)
            self.assertIn("주가 data", workbook.sheetnames)
            self.assertIn("주가", workbook.sheetnames)
            self.assertEqual(workbook["주가 data"]["A1"].value, "거래일")
            self.assertEqual(workbook["주가"]["A13"].value, "거래일")

    def test_price_event_is_not_repeated_for_one_source_event(self) -> None:
        points = [
            PricePoint("2026-08-21", 100.0, None, "https://example.com/price"),
            PricePoint("2026-08-24", 110.0, None, "https://example.com/price"),
            PricePoint("2026-08-25", 121.0, None, "https://example.com/price"),
        ]
        matter = MatterFact("웹 이슈 조사", "계약", "해석", "verified", "event-1", "공식 보도자료", "2026-08-24", "https://example.com/event")
        results = price_event_matters(points, [matter])
        self.assertEqual(len([item for item in results if item.category == "주가 변동·이슈 대조"]), 1)

    def test_gemini_result_requires_known_source_id(self) -> None:
        source = MatterFact("웹 이슈 조사", "공식 사실", "해석", "verified", "official-1", "공식 발표", "2026-01-01", "https://example.com")
        payload = {"items": [
            {"category": "계약", "fact": "근거 있는 사실", "interpretation": "근거 있는 해석", "source_ids": ["official-1"], "verification_status": "verified"},
            {"category": "계약", "fact": "근거 없는 사실", "interpretation": "근거 없는 해석", "source_ids": ["invented"], "verification_status": "verified"},
        ]}
        results = _to_matters(payload, [source])
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0].source_document_id, "official-1")
